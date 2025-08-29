from pprint import pprint
from openpyxl.reader.excel import load_workbook
from datetime import datetime
import pandas as pd

# wb = load_workbook("C:/Users/KNV/PycharmProjects/fifteenth/data.xlsx", data_only=True)
# wb = load_workbook('data.xlsx', data_only=True)
# excel_file = pd.ExcelFile("C:/Users/KNV/PycharmProjects/fifteenth/data.xlsx")
excel_file = pd.ExcelFile("data.xlsx")
# sheets_data = [excel_file.parse(sheet_name, header=None, usecols=[0, 1, 2]) for sheet_name in excel_file.sheet_names]

names_df = {sheet_name: excel_file.parse(sheet_name, header=None, usecols=[0]) for sheet_name in excel_file.sheet_names}
start_dates_df = {sheet_name: excel_file.parse(sheet_name, header=None, usecols=[1]) for sheet_name in excel_file.sheet_names}
finish_dates_df = {sheet_name: excel_file.parse(sheet_name, header=None, usecols=[2]) for sheet_name in excel_file.sheet_names}

start_dates = []
for df in start_dates_df.values():
    dates_list = pd.to_datetime(df.values.ravel(), errors='coerce', format='%d-%m-%Y').to_list()
    start_dates.extend(dates_list)
finish_dates = []
for df in finish_dates_df.values():
    dates_list2 = pd.to_datetime(df.values.ravel(), errors="coerce", format='%d-%m-%Y').to_list()
    finish_dates.extend(dates_list2)

time_deltas = []

print(start_dates)
# print(finish_dates)
